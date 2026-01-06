import datetime
import pandas as pd
from typing import Dict, Set, List, Optional
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont

from core.ports.ranking_report_port import RankingReportPort
from core.ports.storage_port import StoragePort
from core.ports.price_data_port import PriceDataPort
from core.services.high_price_indicator_service import HighPriceIndicatorService
from infra.adapters.excel.excel_formatter import ExcelFormatter
from infra.adapters.excel.excel_sheet_builder import ExcelSheetBuilder


class RankingExcelAdapter(RankingReportPort):
    """순위표를 Excel 형식으로 생성하는 어댑터.

    RankingReportPort 인터페이스의 Excel 구현체입니다.
    ExcelFormatter와 ExcelSheetBuilder 유틸리티를 조합하여 사용합니다.

    Attributes:
        storage (StoragePort): 파일 저장/로드 포트.
        file_path (str): Excel 파일 경로.
    """
    
    TOP_N = 30
    LAYOUT_MAP = {
        'KOSPI_foreigner': {'stock_col': 'E', 'value_col': 'F', 'rank_col': 'D', 'high_price_col': 'G', 'start_row': 5, 'market': 'KOSPI'},
        'KOSPI_institutions': {'stock_col': 'I', 'value_col': 'J', 'rank_col': 'H', 'high_price_col': 'K', 'start_row': 5, 'market': 'KOSPI'},
        'KOSDAQ_foreigner': {'stock_col': 'N', 'value_col': 'O', 'rank_col': 'M', 'high_price_col': 'P', 'start_row': 5, 'market': 'KOSDAQ'},
        'KOSDAQ_institutions': {'stock_col': 'R', 'value_col': 'S', 'rank_col': 'Q', 'high_price_col': 'T', 'start_row': 5, 'market': 'KOSDAQ'},
    }
    # Top 30 기준 Clear Range: 5행부터 34행까지 (30개)
    COLUMNS_TO_AUTOFIT = [chr(i) for i in range(ord('C'), ord('T') + 1)]
    KOREAN_WEEKDAYS = ["월", "화", "수", "목", "금", "토", "일"]
    
    # 기본 템플릿 경로 상수 (StorageRoot 기준)
    DEFAULT_TEMPLATE_PATH = "template/template_일별수급순위정리표.xlsx"
    
    DEFAULT_TEMPLATE_PATH = "template/template_일별수급순위정리표.xlsx"
    
    def __init__(
        self, 
        source_storage: StoragePort, 
        target_storages: List[StoragePort],
        price_port: Optional[PriceDataPort] = None,
        template_file_path: str = None
    ):
        """RankingExcelAdapter 초기화.

        Args:
            source_storage (StoragePort): 파일을 로드할 저장소 (예: GoogleDriveAdapter).
            target_storages (List[StoragePort]): 파일을 저장할 저장소 리스트 (예: [LocalStorageAdapter, GoogleDriveAdapter]).
            price_port (Optional[PriceDataPort]): 가격 데이터 조회 포트 (신고가 지표용).
            template_file_path (str): 템플릿 파일 경로 (Optional).
        """
        self.source_storage = source_storage
        self.target_storages = target_storages
        self.price_port = price_port
        self.file_path = None # update_report 시 결정됨
        self.template_file_path = template_file_path or self.DEFAULT_TEMPLATE_PATH
        
        # 신고가 서비스 초기화 (price_port가 있는 경우에만)
        self.high_price_service = HighPriceIndicatorService(price_port) if price_port else None
        
        print(f"[Adapter:RankingExcel] 초기화 완료 (템플릿: {self.template_file_path}, 신고가: {self.price_port is not None})")
    
    def update_report(
        self,
        report_date: datetime.date,
        data_map: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]]
    ) -> bool:
        """순위표 리포트를 업데이트합니다.

        Args:
            report_date (datetime.date): 리포트 날짜.
            data_map (Dict[str, pd.DataFrame]): 데이터 딕셔너리.
            common_stocks (Dict[str, Set[str]]): 공통 종목 딕셔너리.

        Returns:
            bool: 성공 여부.
        """
        # 동적 파일 경로 설정 ({Year}년/일별수급정리표/{Year}일별수급순위정리표.xlsx)
        year = report_date.year
        self.file_path = f"{year}년/일별수급정리표/{year}일별수급순위정리표.xlsx"
        
        book = self._load_workbook()
        if not book:
            return False
        
        # 이전 순위 파싱 (새 시트 생성 전에 수행)
        previous_rankings = self._parse_previous_rankings(book)
        
        # 신고가 지표 분석 (price_port가 있는 경우에만)
        high_price_indicators = {}
        if self.high_price_service:
            ticker_map = self._create_ticker_map(data_map)
            date_str = report_date.strftime('%Y%m%d')
            high_price_indicators = self.high_price_service.analyze_high_price_indicators(ticker_map, date_str)
        
        # 연속 등장 분석
        streaks = self._analyze_consecutive_streaks(book, report_date)
        
        new_sheet = self._create_new_sheet(book, report_date)
        if not new_sheet:
            return False
        
        self._update_sheet_content(
            new_sheet, 
            report_date, 
            data_map, 
            common_stocks, 
            previous_rankings, 
            high_price_indicators,
            streaks
        )
        
        # 템플릿 시트 제거 (사용자 요청)
        if 'template' in book.sheetnames:
            del book['template']
            print(f"    -> [Adapter:RankingExcel] 'template' 시트 제거 완료")
        
        return self._save_workbook(book)
    
    def _load_workbook(self) -> Workbook | None:
        """워크북을 로드합니다. 파일이 없으면 템플릿을 복사하여 시작합니다."""
        print(f"    -> [Adapter:RankingExcel] 로드 시도 ({self.source_storage.__class__.__name__})...")
        
        # 파일이 존재하는지 확인
        # 파일이 존재하는지 확인
        if not self.source_storage.path_exists(self.file_path):
            print(f"    -> [Adapter:RankingExcel] 파일이 없어 템플릿 복사를 시도합니다: {self.template_file_path}")
            
            # 템플릿 파일 로드 (항상 로컬 파일시스템 사용)
            # source_storage가 Google Drive일 경우에도 템플릿은 로컬에서 읽어서 사용하기 위함
            import os
            template_data = None
            
            # 로컬 경로 찾기 시도 (CWD 기준 또는 output 폴더 기준)
            candidates = [
                self.template_file_path,
                os.path.join("output", self.template_file_path)
            ]
            
            real_template_path = None
            for p in candidates:
                if os.path.exists(p):
                    real_template_path = p
                    break
            
            try:
                if real_template_path:
                    print(f"    -> [Adapter:RankingExcel] 로컬 템플릿 파일 발견: {real_template_path}")
                    with open(real_template_path, 'rb') as f:
                        template_data = f.read()
            except Exception as e:
                print(f"    -> [Adapter:RankingExcel] 🚨 로컬 템플릿 읽기 오류: {e}")

            if template_data:
                # 타겟 경로에 템플릿 저장 (Source Storage에 우선 저장하여 로드 가능하게 함)
                # 주의: 로드는 source_storage에서 하므로, source_storage에 파일이 있어야 함.
                if self.source_storage.put_file(self.file_path, template_data):
                    print(f"    -> [Adapter:RankingExcel] 템플릿 복사 및 업로드 성공")
                else:
                    print(f"    -> [Adapter:RankingExcel] 🚨 템플릿 저장(업로드) 실패")
                    return None
            else:
                print(f"    -> [Adapter:RankingExcel] 🚨 로컬 템플릿 파일을 찾을 수 없습니다: {self.template_file_path}")
                # 템플릿이 없으면 새 파일 생성 로직으로 갈 수도 있지만, 여기서는 실패 처리
                return None

        # 파일 로드
        book = self.source_storage.load_workbook(self.file_path)
        if not book:
            print(f"    -> [Adapter:RankingExcel] 🚨 워크북 로드 실패: {self.file_path}")
            return None
            
        return book

    def _analyze_consecutive_streaks(self, book: Workbook, current_date: datetime.date) -> Dict[str, Dict[str, int]]:
        """과거 시트들을 분석하여 연속 등장 횟수를 계산합니다."""
        streaks = {}
        
        # 분석할 시트 목록 가져오기 (template 제외, 최신순 정렬)
        valid_sheets = [s for s in book.worksheets if s.title != 'template']
        # 이미 날짜순으로 되어 있다고 가정 (append 방식이므로)
        # 역순으로 순회 (최신 -> 과거)
        
        if not valid_sheets:
            return streaks

        print(f"    -> [Adapter:RankingExcel] 연속 등장 분석 시작 ({len(valid_sheets)}개 시트)")

        # 각 섹션별로 분석
        for key, layout in self.LAYOUT_MAP.items():
            section_streaks = {}
            stock_col = layout['stock_col']
            start_row = layout['start_row']
            
            # 과거 10일치 정도만 조회 (최적화)
            history_dfs = []
            
            # 최근 시트부터 역순으로 
            # (주의: 현재 리포트 날짜의 시트는 아직 생성 안됨. valid_sheets는 모두 과거)
            for sheet in reversed(valid_sheets):
                sheet_stocks = set()
                for i in range(self.TOP_N):
                    row = start_row + i
                    val = sheet[f"{stock_col}{row}"].value
                    # (쌍) 등 제거
                    if val and isinstance(val, str):
                        clean_name = val.split(' (')[0]
                        sheet_stocks.add(clean_name)
                history_dfs.append(sheet_stocks)
                if len(history_dfs) >= 5: # 5일 이상이면 충분 (5+가 최대 등급이므로)
                    break
            
            # 여기서 스트릭을 미리 계산할 수는 없음 (오늘 데이터가 뭔지 모르므로)
            # 대신 과거 데이터를 리스트 형태로 저장해두고, 나중에 오늘 데이터와 비교?
            # 아니함수 구조상 오늘 데이터(data_map)를 여기서 알 수 없으므로,
            # '오늘 등장한다면 며칠 연속인가'를 계산해두는게 좋음.
            
            # 즉, "오늘 만약 A가 등장한다면, 어제도 A가 있었는가? 그 엊그제도?" 를 미리 카운트.
            # 모든 과거에 등장했던 종목에 대해 현재 기준 연속성을 미리 계산.
            
            potential_streaks = {} 
            # history_dfs[0] = 어제, [1] = 엊그제...
            
            if not history_dfs:
                streaks[key] = {}
                continue

            # 어제 등장했던 모든 종목에 대해 조사
            yesterday_stocks = history_dfs[0]
            for stock in yesterday_stocks:
                count = 1 # 어제 등장했으므로 최소 1 (오늘 등장시 2가 됨)
                
                # 엊그제부터 체크
                for past_stocks in history_dfs[1:]:
                    if stock in past_stocks:
                        count += 1
                    else:
                        break # 연속 끊김
                
                potential_streaks[stock] = count
            
            streaks[key] = potential_streaks
            
        return streaks

    def _parse_previous_rankings(self, book: Workbook) -> Dict[str, Dict[str, int]]:
        """마지막 시트에서 이전 순위를 파싱합니다."""
        rankings = {}
        
        # 시트가 없거나 'template' 하나뿐이면 이전 데이터 없음
        if not book.worksheets or (len(book.worksheets) == 1 and 'template' in book.sheetnames):
            return rankings
            
        # 마지막 시트 (직전 거래일)
        last_sheet = book.worksheets[-1]
        print(f"    -> [Adapter:RankingExcel] 이전 순위 데이터 파싱 (Source: {last_sheet.title})")
        
        for key, layout in self.LAYOUT_MAP.items():
            section_ranks = {}
            stock_col = layout['stock_col']
            start_row = layout['start_row']
            
            # Top N 만큼 순회
            for i in range(self.TOP_N):
                row = start_row + i
                stock_cell = last_sheet[f"{stock_col}{row}"]
                stock_name = stock_cell.value
                
                if stock_name and isinstance(stock_name, str):
                    section_ranks[stock_name] = i + 1  # 1-based rank
            
            rankings[key] = section_ranks
            
        return rankings
    
    def _create_ticker_map(self, data_map: Dict[str, pd.DataFrame]) -> Dict[str, str]:
        """데이터맵에서 종목명-종목코드 매핑을 생성합니다.
        
        Args:
            data_map (Dict[str, pd.DataFrame]): 데이터 딕셔너리 (종목코드 컬럼 포함).
            
        Returns:
            Dict[str, str]: 종목명 -> 종목코드 매핑.
        """
        ticker_map = {}
        
        for key, df in data_map.items():
            if df is None or df.empty:
                continue
            
            # DataFrame에 종목코드와 종목명 컬럼이 있는지 확인
            if '종목코드' in df.columns and '종목명' in df.columns:
                for _, row in df.iterrows():
                    stock_name = row['종목명']
                    ticker = row['종목코드']
                    
                    if stock_name and ticker:
                        # 종목코드가 숫자로 들어오는 경우 6자리 문자열로 변환 (예: 5930 -> "005930")
                        if isinstance(ticker, (int, float)):
                            ticker = f"{int(ticker):06d}"
                        else:
                            ticker = str(ticker).strip().zfill(6)
                            
                        ticker_map[stock_name] = ticker
        
        print(f"    -> [Adapter:RankingExcel] 종목코드 매핑 생성 완료 ({len(ticker_map)}개)")
        return ticker_map
    
    def _create_new_sheet(self, book: Workbook, report_date: datetime.date) -> Worksheet | None:
        """새로운 시트를 생성합니다 (템플릿 시트 복제)."""
        try:
            sheet_name = report_date.strftime('%m%d')
            
            # 이미 시트가 있으면 삭제
            if sheet_name in book.sheetnames:
                del book[sheet_name]
            
            # 복제 소스 시트 결정 ('template' 시트 우선)
            if 'template' in book.sheetnames:
                source_sheet = book['template']
                print(f"    -> [Adapter:RankingExcel] 'template' 시트 복제 사용")
            else:
                source_sheet = book.worksheets[-1]
                print(f"    -> [Adapter:RankingExcel] 'template' 시트가 없어 마지막 시트 복제 사용")
            
            new_sheet = book.copy_worksheet(source_sheet)
            new_sheet.title = sheet_name
            
            
            
            # 시트 보호 해제 (편집 가능하도록 설정)
            if new_sheet.protection:
                new_sheet.protection.sheet = False
            
            print(f"    -> [Adapter:RankingExcel] '{sheet_name}' 시트 생성 완료 (보호 해제됨)")
            return new_sheet
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"    -> [Adapter:RankingExcel] 🚨 시트 생성 실패: {e}")
            return None
    
    def _update_sheet_content(
        self,
        sheet: Worksheet,
        report_date: datetime.date,
        data_map: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]],
        previous_rankings: Dict[str, Dict[str, int]],
        high_price_indicators: Dict[str, Dict[str, Optional[str]]] = None,
        streaks: Dict[str, Dict[str, int]] = None
    ):
        """시트 내용을 업데이트합니다."""
        self._update_headers(sheet, report_date)
        self._clear_data_area(sheet)
        self._paste_data_and_apply_format(
            sheet, data_map, common_stocks, previous_rankings, 
            high_price_indicators or {}, streaks or {}
        )
        self._apply_autofit(sheet)
    
    def _update_headers(self, sheet: Worksheet, report_date: datetime.date):
        """헤더를 업데이트합니다.
        
        A3: 월 (예: "11 月")
        A5: 일 (예: "21 日")
        B5: 요일 (예: "금")
        """
        sheet['A3'] = f"{report_date.month} 月"
        sheet['A5'] = f"{report_date.day} 日"
        sheet['B5'] = self.KOREAN_WEEKDAYS[report_date.weekday()]
    
    def _clear_data_area(self, sheet: Worksheet):
        """데이터 영역을 초기화합니다. (레이아웃에 정의된 데이터 컬럼만 초기화)"""
        # 레이아웃에 정의된 컬럼만 선택적으로 초기화하여 J열(순위 등)은 유지
        clear_limit = self.TOP_N + 5  # TOP_N 보다 조금 더 넉넉하게 초기화
        
        for key, layout in self.LAYOUT_MAP.items():
            # Stock, Value, Rank, HighPrice 컬럼 초기화
            cols_to_clear = [layout['stock_col'], layout['value_col']]
            if 'rank_col' in layout:
                cols_to_clear.append(layout['rank_col'])
            if 'high_price_col' in layout:
                cols_to_clear.append(layout['high_price_col'])
            
            start_row = layout['start_row']
            
            for col in cols_to_clear:
                for i in range(clear_limit):
                    row_idx = start_row + i
                    cell = sheet[f"{col}{row_idx}"]
                    cell.value = None
                    # Rank 컬럼은 서식을 유지하거나 재설정하는데, RichText를 쓰려면 초기화가 나음
                    cell.fill = PatternFill()
    
    def _paste_data_and_apply_format(
        self,
        sheet: Worksheet,
        data_map: Dict[str, pd.DataFrame],
        common_stocks: Dict[str, Set[str]],
        previous_rankings: Dict[str, Dict[str, int]],
        high_price_indicators: Dict[str, Dict[str, Optional[str]]],
        streaks: Dict[str, Dict[str, int]]
    ):
        """데이터를 붙여넣고 서식을 적용합니다."""
        for key, layout in self.LAYOUT_MAP.items():
            df = data_map.get(key)
            if df is None or df.empty:
                continue
            
            # DataFrame 붙여넣기 (종목명, 금액)
            pasted_count = ExcelSheetBuilder.paste_ranking_data(sheet, df, layout, self.TOP_N)
            
            # 순위 변동 기입 (Rich Text)
            prev_section_ranks = previous_rankings.get(key, {})
            rank_col = layout.get('rank_col')
            stock_col = layout['stock_col']
            start_row = layout['start_row']
            section_streaks = streaks.get(key, {})
            
            for i in range(pasted_count):
                row = start_row + i
                stock_cell = sheet[f"{stock_col}{row}"]
                stock_name = stock_cell.value
                
                # 순위 변동
                if rank_col:
                    current_rank = i + 1
                    prev_rank = prev_section_ranks.get(stock_name)
                    diff = None
                    if prev_rank:
                        diff = prev_rank - current_rank
                    self._write_rank_change(sheet, rank_col, row, diff)
                
                # 연속 등장 하이라이트 적용
                if stock_name:
                    clean_name = stock_name.replace(' (쌍)', '')
                    # 어제까지의 연속 횟수 + 오늘(1) = 현재 연속 횟수
                    # section_streaks에는 '어제까지의 연속 횟수'가 들어있음
                    streak_count = section_streaks.get(clean_name, 0) + 1
                    
                    streak_color = None
                    if streak_count >= 5:
                        streak_color = 'red'
                    elif streak_count == 4:
                        streak_color = 'orange'
                    elif streak_count == 3:
                        streak_color = 'yellow'
                    elif streak_count == 2:
                        streak_color = 'green'
                    
                    if streak_color and streak_color in ExcelFormatter.COLORS:
                        fill = PatternFill(
                            start_color=ExcelFormatter.COLORS[streak_color],
                            end_color=ExcelFormatter.COLORS[streak_color],
                            fill_type="solid"
                        )
                        stock_cell.fill = fill
                
            
            # 신고가 지표 표시
            high_price_col = layout.get('high_price_col')
            if high_price_col and high_price_indicators:
                for i in range(pasted_count):
                    row = start_row + i
                    stock_name = sheet[f"{stock_col}{row}"].value
                    
                    # (쌍) 표시가 있으면 제거하여 비교
                    clean_stock_name = stock_name.replace(' (쌍)', '') if stock_name else None
                    
                    if clean_stock_name and clean_stock_name in high_price_indicators:
                        indicator = high_price_indicators[clean_stock_name]
                        text = indicator.get('text')
                        color = indicator.get('color')
                        
                        if text and color:
                            self._write_high_price_indicator(sheet, high_price_col, row, text, color)
            
            # 공통 종목에 (쌍) 표시 추가
            market = layout['market']
            if market in common_stocks:
                stock_col = layout['stock_col']
                start_row = layout['start_row']
                for i in range(pasted_count):
                    row = start_row + i
                    stock_cell = sheet[f"{stock_col}{row}"]
                    stock_name = stock_cell.value
                    if stock_name and stock_name in common_stocks[market]: # (쌍) 중복 방지를 위해 값 체크는 단순하게
                        # 이미 (쌍)이 붙어있을 수 있음(위 하이라이트 로직에서 건드리지 않음)
                        # 하지만 확실히 하기 위해 clean_name 비교
                        clean = str(stock_name).replace(' (쌍)', '')
                        if clean in common_stocks[market]:
                             stock_cell.value = f"{clean} (쌍)"
            
            ExcelSheetBuilder.clear_ranking_remaining_rows(sheet, layout, pasted_count, self.TOP_N)

            
    def _write_rank_change(self, sheet: Worksheet, col: str, row: int, diff: int | None):
        """순위 변동을 Rich Text로 기입합니다."""
        cell = sheet[f"{col}{row}"]
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if diff is None:  # New Entry
            cell.value = "✨"
            return

        abs_diff = abs(diff)
        
        if diff >= 15:
            # 급상승 (Big Red Triangle + Black Number)
            # Bold Red Triangle
            symbol = TextBlock(InlineFont(color='FF0000', sz=14, b=True), "▲")
            number = TextBlock(InlineFont(color='000000', sz=11), str(abs_diff))
            cell.value = CellRichText([symbol, number])
            
        elif diff > 0:
            # 상승 (Red Triangle + Black Number)
            symbol = TextBlock(InlineFont(color='FF0000'), "▲")
            number = TextBlock(InlineFont(color='000000'), str(abs_diff))
            cell.value = CellRichText([symbol, number])
            
        elif diff < 0:
            # 하락 (Blue Triangle + Black Number)
            symbol = TextBlock(InlineFont(color='0000FF'), "▼")
            number = TextBlock(InlineFont(color='000000'), str(abs_diff))
            cell.value = CellRichText([symbol, number])
            
        else:
            # 유지
            cell.value = "-"
    
    def _write_high_price_indicator(self, sheet: Worksheet, col: str, row: int, text: str, color_key: str):
        """신고가 지표를 셀에 기입합니다.
        
        Args:
            sheet (Worksheet): 워크시트.
            col (str): 열 문자.
            row (int): 행 번호.
            text (str): 표시 텍스트.
            color_key (str): 색상 키 (ExcelFormatter.COLORS).
        """
        cell = sheet[f"{col}{row}"]
        cell.value = text
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 배경색 적용
        if color_key in ExcelFormatter.COLORS:
            fill = PatternFill(
                start_color=ExcelFormatter.COLORS[color_key],
                end_color=ExcelFormatter.COLORS[color_key],
                fill_type="solid"
            )
            cell.fill = fill

            
    def _apply_autofit(self, sheet: Worksheet):
        """열 너비를 자동 조정합니다."""
        for col in self.COLUMNS_TO_AUTOFIT:
            sheet.column_dimensions[col].bestFit = True
            sheet.column_dimensions[col].auto_size = True
    
    def _save_workbook(self, book: Workbook) -> bool:
        """워크북을 저장합니다 (Target Storages 사용)."""
        all_success = True
        for storage in self.target_storages:
            success = storage.save_workbook(book, self.file_path)
            if success:
                print(f"    -> [Adapter:RankingExcel] ✅ {storage.__class__.__name__} 순위표 저장 완료")
            else:
                all_success = False
        return all_success
